import os
from helper import clocks_file_path, excel_file_path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# atver vai izveido Excel failu
def open_or_create_workbook(filename):
    try:
        workbook = load_workbook(filename, data_only=True)
    except FileNotFoundError:
        workbook = Workbook()
    return workbook

def generate_excel():
    # atver vai izveido Excel failu
    workbook = open_or_create_workbook(excel_file_path)

    # Pievieno jaunu lapu
    previous_month_date = datetime.today().replace(day = 1) - timedelta(days = 1)
    sheet_name = previous_month_date.strftime('%b%y')
    if sheet_name not in workbook.sheetnames:
        if len(workbook.sheetnames) == 1 and workbook.sheetnames[0] == 'Sheet': # Fails bija tikko izveidots (tikai viena lapa ar nosaukumu 'Sheet')
            workbook.active.title = sheet_name # tad jāpārsauc to lapu
        else:
            workbook.create_sheet(title=sheet_name) # citādi jāizveido jauno

    sheet = workbook[sheet_name]

    # galvenes (viena tukša kolona)
    #            B1         C1          D3      E1      F1
    headers = ["Date", "Clock In", "Clock Out", "", "Hours"]
    for col, header in enumerate(headers, 2):  # sākot ar otro kolonu
        if not header:
            continue
        sheet[f"{get_column_letter(col)}1"] = header
        sheet[f"{get_column_letter(col)}1"].style = 'Check Cell'

    # Nolasa datus no faila clock.txt un kopē tos Excel lapā
    open(clocks_file_path, 'a') # ja nav, izveido
    with open(clocks_file_path, 'r') as file:
        clocks = file.readlines()
        for row_index, line in enumerate(clocks, 2):  # sākot ar Excel otro rindu
            # date var būt tukšs
            *date, clock_in, clock_out = line.strip().split('\t')
            if not clock_in or not clock_out: # defekta rinda
                continue
            sheet[f"B{row_index}"] = date[0] if date else "" # var būt tukša
            sheet[f"C{row_index}"] = clock_in
            sheet[f"C{row_index}"].number_format = 'hh:mm'
            sheet[f"D{row_index}"] = clock_out
            sheet[f"D{row_index}"].number_format = 'hh:mm'

            # Formula for hours calculation
            sheet[f"F{row_index}"] = f"=D{row_index}-C{row_index}"
            sheet[f"F{row_index}"].number_format = 'hh:mm'

            # hh, _ = 'hh:mm'.split(':')
            clock_in_hour, _ = clock_in.split(':')
            # hh, mm = 'hh:mm'.split(':')
            clock_out_hour, clock_out_min = clock_out.split(':')
            if (int(clock_in_hour) > int(clock_out_hour)): # ja Clock In > Clock Out (negatīvs nostrādātu stundu skaits), tād Clock Out jāpalielinā uz 24h
                sheet[f"D{row_index}"] = f"{int(clock_out_hour)+24}:{clock_out_min}"

    # beigās sasummētas stundas
    if clocks:
        sheet[f"F{row_index+1}"] = f'=SUM(F2:F{row_index})'
        sheet[f"F{row_index+1}"].style = '40 % - Accent4'
        sheet[f"F{row_index+1}"].number_format = '[h]:mm'

    workbook.save(excel_file_path)
    os.remove(clocks_file_path) # izdēsa TXT failu

if __name__ == '__main__':
    generate_excel()